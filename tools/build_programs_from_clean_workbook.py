from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

DATE_RE = re.compile(r'([A-Za-z]{3,9})\s+([A-Za-z]{3})\s+(\d{1,2}),\s*(\d{4})')
TIME_RE = re.compile(r'^(\d{1,2}):(\d{2})$')
EPISODE_RE = re.compile(r'^(#.*)$')
MONTH_NAMES = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
}

@dataclass
class ProgramRow:
    date: str
    day: str
    start: str
    title: str
    episode: str
    durationMinutes: int
    seasonStart: bool = False


def clean_text(value: object) -> str:
    if value is None:
        return ''
    text = str(value).replace('\r', '\n')
    lines = [ln.strip() for ln in text.split('\n')]
    lines = [ln for ln in lines if ln]
    return '\n'.join(lines)


def parse_day_header(value: object) -> Optional[Tuple[date, str]]:
    text = clean_text(value)
    if not text:
        return None
    m = DATE_RE.search(text.replace('\n', ' '))
    if not m:
        return None
    weekday, mon_abbr, day_num, year_num = m.groups()
    month_num = MONTH_NAMES.get(mon_abbr)
    if month_num is None:
        return None
    dt = date(int(year_num), month_num, int(day_num))
    return dt, weekday


def parse_time_value(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%H:%M')
    if isinstance(value, time):
        return value.strftime('%H:%M')
    text = clean_text(value)
    if not text:
        return None
    m = TIME_RE.match(text)
    if not m:
        return None
    hh, mm = m.groups()
    return f'{int(hh):02d}:{mm}'


def parse_program_cell(value: object) -> Optional[Tuple[str, str, bool]]:
    text = clean_text(value)
    if not text:
        return None
    lines = text.split('\n')
    title_lines: List[str] = []
    episode = ''
    for ln in lines:
        if EPISODE_RE.match(ln):
            episode = ln
            break
        title_lines.append(ln)
    title = ' '.join(title_lines).strip()
    if not title and not episode:
        return None
    season_start = False
    if episode:
        # Keep it practical. Many season/series starts are marked by episode code ending in 01 or #0101 style.
        digits = re.sub(r'[^0-9]', '', episode)
        if digits.endswith('01'):
            season_start = True
    return title, episode, season_start


def infer_duration_minutes(starts: List[int], idx: int) -> int:
    current = starts[idx]
    if idx + 1 < len(starts):
        nxt = starts[idx + 1]
        delta = max(30, nxt - current)
        return delta
    return max(30, 24 * 60 - current)


def parse_week_sheet(ws, month_filter: Optional[str]) -> List[ProgramRow]:
    rows = list(ws.iter_rows(values_only=True))
    header_row_index = None
    date_cols: Dict[int, Tuple[date, str]] = {}
    for ridx, row in enumerate(rows[:8]):
        for cidx, val in enumerate(row):
            parsed = parse_day_header(val)
            if parsed:
                date_cols[cidx] = parsed
        if date_cols:
            header_row_index = ridx
            break
    if header_row_index is None:
        return []

    extracted: Dict[str, List[Tuple[int, str, str, bool]]] = defaultdict(list)
    day_names: Dict[str, str] = {}

    for row in rows[header_row_index + 1:]:
        if not row:
            continue
        row_time = parse_time_value(row[0] if len(row) > 0 else None)
        if row_time is None:
            continue
        start_minutes = int(row_time[:2]) * 60 + int(row_time[3:])
        for cidx, (dt, weekday) in date_cols.items():
            if cidx >= len(row):
                continue
            if month_filter and dt.strftime('%Y-%m') != month_filter:
                continue
            parsed = parse_program_cell(row[cidx])
            if not parsed:
                continue
            title, episode, season_start = parsed
            iso_date = dt.isoformat()
            day_names[iso_date] = weekday
            extracted[iso_date].append((start_minutes, title, episode, season_start))

    out: List[ProgramRow] = []
    for iso_date, starts in sorted(extracted.items()):
        starts.sort(key=lambda x: x[0])
        minute_list = [s[0] for s in starts]
        for i, (start_minutes, title, episode, season_start) in enumerate(starts):
            duration = infer_duration_minutes(minute_list, i)
            hh, mm = divmod(start_minutes, 60)
            out.append(ProgramRow(
                date=iso_date,
                day=day_names[iso_date],
                start=f'{hh:02d}:{mm:02d}',
                title=title,
                episode=episode,
                durationMinutes=duration,
                seasonStart=season_start,
            ))
    return out


def dedupe_programs(rows: List[ProgramRow]) -> List[ProgramRow]:
    seen = set()
    out = []
    for row in sorted(rows, key=lambda r: (r.date, r.start, r.title, r.episode)):
        key = (row.date, row.start, row.title, row.episode)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def validate(rows: List[ProgramRow]) -> List[str]:
    problems: List[str] = []
    by_date = defaultdict(list)
    for r in rows:
        by_date[r.date].append(r)
    for iso_date, items in by_date.items():
        items.sort(key=lambda r: r.start)
        occupied = set()
        for r in items:
            start = int(r.start[:2]) * 60 + int(r.start[3:])
            slots = max(1, round(r.durationMinutes / 30))
            for i in range(slots):
                key = start + i * 30
                if key in occupied:
                    problems.append(f'Overlap on {iso_date} at {r.start}: {r.title}')
                occupied.add(key)
        if not occupied:
            continue
        mins = sorted(occupied)
        for a, b in zip(mins, mins[1:]):
            if b - a > 30:
                problems.append(f'Gap on {iso_date} between {a//60:02d}:{a%60:02d} and {b//60:02d}:{b%60:02d}')
                break
    return problems


def write_js(rows: List[ProgramRow], channel: str, month: str, output_path: Path) -> None:
    serializable = [r.__dict__ for r in rows]
    js = (
        f"window.MONTH_CONFIG = {{ channel: {json.dumps(channel)}, month: {json.dumps(month)} }};\n"
        f"window.MONTH_PROGRAMS = {json.dumps(serializable, indent=2)};\n"
    )
    output_path.write_text(js, encoding='utf-8')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Build external month program data from a clean workbook.')
    parser.add_argument('workbook', help='Path to the clean schedule workbook (.xlsx)')
    parser.add_argument('--month', default='2026-05', help='Month filter in YYYY-MM format')
    parser.add_argument('--channel', default='wnmu1hd', help='Channel slug for output metadata')
    parser.add_argument('--output', default='', help='Output JS path. Defaults beside workbook as <channel>-<month>-programs.js')
    args = parser.parse_args()

    wb = load_workbook(args.workbook, data_only=True)
    rows: List[ProgramRow] = []
    for ws in wb.worksheets:
        if 'Grid' not in ws.title:
            continue
        rows.extend(parse_week_sheet(ws, args.month))
    rows = dedupe_programs(rows)
    problems = validate(rows)

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = Path(args.workbook).with_name(f'{args.channel}-{args.month}-programs.js')
    write_js(rows, args.channel, args.month, output_path)

    print(f'Wrote {len(rows)} programs to {output_path}')
    if problems:
        print('Validation notes:')
        for p in problems[:50]:
            print(' -', p)
    else:
        print('No overlaps or obvious gaps detected in populated days.')
